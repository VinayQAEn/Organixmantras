import openpyxl
from openpyxl.styles import PatternFill


def get_rowcount(file,Sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[Sheetname]
    return (sheet.max_row)

def get_column(file,Sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[Sheetname]
    return sheet.max_column

def readdata(file,Sheetname,rowno,columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[Sheetname]
    return sheet.cell(rowno,columno).value

def writedata(file,Sheetname,rowno,columnno,DatA):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[Sheetname]
    sheet.cell(rowno,columnno).value = DatA
    workbook.save(file)

def green_fill(file, sheetname, rowno, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    greenfill = PatternFill(start_color="60b212", end_color="60b212", fill_type='solid')
    sheet.cell(row=rowno, column=columnno).fill = greenfill
    workbook.save(file)

def red_fill(file, sheetname, rowno, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    redfill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type='solid')
    sheet.cell(row=rowno, column=columnno).fill = redfill
    workbook.save(file)

